"""
Db Management: delete, import from excel, export to excel.
"""
import logging

from django.db import transaction
from django.utils.text import format_lazy
from django.utils.translation import gettext_lazy as _


logger = logging.getLogger('ohmi_audit')


class DbManagement:
    """
    This class is responsible for managing the database, including deleting records, importing data
    from Excel files, and exporting data to Excel files.
    """
    @staticmethod
    def delete_database():
        """
        Deletes all records from the application tables (Audit, Auditor, Customer)
        and removes any associated media files (images / uploaded files).
        Non-superuser, non-staff AppUsers are also removed.

        All deletions run inside a single atomic transaction so that any
        unexpected error will roll back every change made so far.

        Returns a translated success message on completion.
        Raises an exception on failure (the view layer catches it).
        """
        # Lazy import to avoid circular-import issues at module load time.
        from ohmi_audit.main_app.models import Audit, Auditor, Customer, AppUser

        try:
            with transaction.atomic():
                # ------------------------------------------------------------------
                # Step 1 – remove media files that belong to Audit records BEFORE
                # the database rows are deleted (we still need the field values).
                # ------------------------------------------------------------------
                for audit in Audit.objects.exclude(image='').exclude(image__isnull=True):
                    try:
                        audit.image.delete(save=False)   # deletes file from storage
                    except Exception as file_err:
                        logger.warning("Could not delete image for Audit pk=%s: %s", audit.pk, file_err)

                for audit in Audit.objects.exclude(file='').exclude(file__isnull=True):
                    try:
                        audit.file.delete(save=False)    # deletes file from storage
                    except Exception as file_err:
                        logger.warning("Could not delete file for Audit pk=%s: %s", audit.pk, file_err)

                # ------------------------------------------------------------------
                # Step 2 – delete database rows.
                # Order matters when FK constraints are present; delete dependants first.
                # ------------------------------------------------------------------
                audit_count, _d    = Audit.objects.all().delete()
                auditor_count, _d  = Auditor.objects.all().delete()
                customer_count, _d = Customer.objects.all().delete()

                # Delete non-superuser / non-staff app users.
                user_count, _d = AppUser.objects.filter(is_superuser=False, is_staff=False).delete()

                logger.info(
                    "DB delete completed – Audits: %d, Auditors: %d, Customers: %d, Users: %d",
                    audit_count, auditor_count, customer_count, user_count,
                )

            return format_lazy(
                _(
                    "Database deleted successfully. "
                    "Removed {a} audit(s), {au} auditor(s), {c} customer(s), {u} user(s)."
                ),
                a=audit_count,
                au=auditor_count,
                c=customer_count,
                u=user_count,
            )

        except Exception as exc:
            logger.error("delete_database failed: %s", exc, exc_info=True)
            raise

    # ------------------------------------------------------------------
    # Private helpers for import_from_excel
    # ------------------------------------------------------------------
    @staticmethod
    def _load_worksheet(uploaded_file):
        """
        Step 1 – Open the uploaded .xlsx file and return the active worksheet.

        `uploaded_file` is Django's InMemoryUploadedFile / TemporaryUploadedFile,
        which openpyxl can read directly as a file-like object.
        `read_only=True` keeps memory usage low for large files.
        `data_only=True` reads cell values instead of formulas.
        """
        import openpyxl
        wb = openpyxl.load_workbook(uploaded_file, read_only=True, data_only=True)
        return wb.active

    @staticmethod
    def _map_headers(header_row):
        """
        Step 2 – Build a dict of { normalised_header: column_index (0-based) }
        from the first row of the worksheet.

        Normalising (strip + lowercase) makes the mapping tolerant of extra
        whitespace in column titles (e.g. 'Company ID ' → 'company id').
        """
        return {
            str(cell).strip().lower(): idx
            for idx, cell in enumerate(header_row)
            if cell is not None
        }

    @staticmethod
    def _build_customer_record(row, header_map):
        """
        Step 3 – Convert one worksheet row into a Customer field dict.

        COLUMN_MAP ties each normalised Excel header to:
          - the matching Customer model field name
          - a type converter (str / int) applied to the raw cell value

        Returns None for blank rows (they are silently skipped).
        Raises ValueError with a descriptive message for bad data.
        """
        # --- Edit this dict when the Excel template changes ---
        COLUMN_MAP = {
            'year':           ('year',           int),   # A – year
            'bg vor.nr.':     ('BG_Vor_Nr',       str),   # B – unique identifier
            'unternehmen-bg': ('company_name_bg', str),   # C – Bulgarian company name
            'unternehmen-en': ('company_name_en', str),   # D – English company name
            'company id':     ('company_id',      int),   # E – numeric company ID
            'vat':            ('VAT_number',       str),   # F – VAT number
        }

        values = list(row)

        # Skip rows where every cell is empty
        if all(v is None or str(v).strip() == '' for v in values):
            return None

        record = {}
        for header_key, (field_name, converter) in COLUMN_MAP.items():
            col_idx = header_map.get(header_key)

            # Guard: required column missing from the file
            if col_idx is None:
                raise ValueError(f"Required column '{header_key}' not found in the Excel file")

            raw = values[col_idx] if col_idx < len(values) else None

            # Guard: required field is empty in this row
            if raw is None or str(raw).strip() == '':
                raise ValueError(f"Empty value for required field '{field_name}'")

            # Convert to the correct Python type
            try:
                record[field_name] = converter(str(raw).strip())
            except (ValueError, TypeError) as exc:
                raise ValueError(f"Invalid value '{raw}' for field '{field_name}': {exc}")

        return record

    # ------------------------------------------------------------------
    # Public method
    # ------------------------------------------------------------------
    @staticmethod
    def import_from_excel(uploaded_file):
        """
        Reads an .xlsx file uploaded through the form and upserts Customer
        records into the database.

        Strategy – update_or_create keyed on BG_Vor_Nr (the unique identifier):
          - first import  → creates new rows
          - re-import     → updates existing rows instead of raising IntegrityError

        Row-level errors are logged and skipped so one bad row never blocks
        the rest of the file from being imported.

        Returns a translated summary message with created / updated / skipped counts.
        """
        from ohmi_audit.main_app.models import Customer

        # ----------------------------------------------------------------
        # Step 1 – load the workbook
        # ----------------------------------------------------------------
        ws = DbManagement._load_worksheet(uploaded_file)
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            raise ValueError("The uploaded Excel file is empty.")

        # ----------------------------------------------------------------
        # Step 2 – map header names → column indices
        # ----------------------------------------------------------------
        header_map = DbManagement._map_headers(rows[0])
        logger.info("Excel headers detected: %s", header_map)

        # ----------------------------------------------------------------
        # Step 3 – process every data row (starting at row index 1 = Excel row 2)
        # ----------------------------------------------------------------
        created_count = 0
        updated_count = 0
        skipped_count = 0   # blank rows
        error_count   = 0

        with transaction.atomic():
            for row_num, row in enumerate(rows[1:], start=2):
                try:
                    record = DbManagement._build_customer_record(row, header_map)

                    # Blank row – skip silently
                    if record is None:
                        skipped_count += 1
                        continue

                    # Upsert: look up by unique BG_Vor_Nr, update everything else
                    lookup   = {'BG_Vor_Nr': record['BG_Vor_Nr']}
                    defaults = {k: v for k, v in record.items() if k != 'BG_Vor_Nr'}

                    _obj, created = Customer.objects.update_or_create(
                        **lookup,
                        defaults=defaults,
                    )
                    if created:
                        created_count += 1
                        logger.info("Row %d – created Customer '%s'", row_num, record['BG_Vor_Nr'])
                    else:
                        updated_count += 1
                        logger.info("Row %d – updated Customer '%s'", row_num, record['BG_Vor_Nr'])

                except Exception as row_err:
                    error_count += 1
                    logger.warning("import_from_excel – row %d skipped: %s", row_num, row_err)

        logger.info(
            "import_from_excel done – created: %d, updated: %d, blank: %d, errors: %d",
            created_count, updated_count, skipped_count, error_count,
        )

        return format_lazy(
            _(
                "Import completed. "
                "Created: {c}, Updated: {u}, Skipped (blank): {s}, Errors: {e}."
            ),
            c=created_count,
            u=updated_count,
            s=skipped_count,
            e=error_count,
        )

    @staticmethod
    def _build_export_workbook(customers):
        """
        Step 1 – Build an openpyxl Workbook from a Customer queryset.

        Column order mirrors the import template exactly so the exported file
        can be re-imported without any modifications.
        Headers are bolded and column widths are auto-fitted for readability.
        """
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Customers'

        # Header row – must match the COLUMN_MAP keys in _build_customer_record
        headers = ['year', 'BG Vor.Nr.', 'Unternehmen-bg', 'Unternehmen-en', 'Company ID', 'VAT']
        ws.append(headers)

        # Bold the header row to match the import file appearance
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Data rows – one per Customer, in the same column order as the headers
        for customer in customers:
            ws.append([
                customer.year,
                customer.BG_Vor_Nr,
                customer.company_name_bg,
                customer.company_name_en,
                customer.company_id,
                customer.VAT_number,
            ])

        # Auto-fit column widths based on the longest value in each column
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 14)

        return wb

    @staticmethod
    def export_to_excel():
        """
        Exports all Customer records to an .xlsx file that mirrors the import
        template format, then returns a file-download HttpResponse.

        The view must return this response directly (not pass it as a template
        context variable) so that the browser triggers an immediate file download.

        Steps:
          1. Fetch all Customer rows ordered by id.
          2. Build the workbook via _build_export_workbook().
          3. Serialise the workbook to an in-memory BytesIO buffer.
          4. Wrap the buffer in an HttpResponse with the correct MIME type and
             Content-Disposition header so the browser saves it as a .xlsx file.
        """
        import io
        from datetime import datetime
        from django.http import HttpResponse
        from ohmi_audit.main_app.models import Customer

        # ----------------------------------------------------------------
        # Step 1 – fetch all customers
        # ----------------------------------------------------------------
        customers = Customer.objects.all().order_by('id')
        count = customers.count()

        # ----------------------------------------------------------------
        # Step 2 – build the workbook
        # ----------------------------------------------------------------
        wb = DbManagement._build_export_workbook(customers)

        # ----------------------------------------------------------------
        # Step 3 – serialise to an in-memory buffer (no temp file on disk)
        # ----------------------------------------------------------------
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # ----------------------------------------------------------------
        # Step 4 – build the download response
        # ----------------------------------------------------------------
        filename = f"customers_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        logger.info("export_to_excel – exported %d customer(s) as '%s'", count, filename)
        return response
